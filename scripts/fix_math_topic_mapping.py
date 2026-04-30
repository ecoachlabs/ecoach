from __future__ import annotations

from pathlib import Path

import openpyxl


SOURCE_PATH = Path(r"C:\Users\surfaceSudio\Downloads\mathematics_topic_mapping.xlsx")
OUTPUT_DIR = Path(r"C:\Users\surfaceSudio\OneDrive\ecoach\outputs")
OUTPUT_PATH = OUTPUT_DIR / "mathematics_topic_mapping_corrected.xlsx"


TOPIC_MAP: dict[str, tuple[str, str]] = {
    "B7.1.1.1": ("Place Value and Rounding", "Place Value & Rounding"),
    "B7.1.2.1": ("Mental Mathematics", "Mental Math"),
    "B7.1.2.2": ("Whole Number and Decimal Operations", "Whole Numbers & Decimals"),
    "B7.1.2.3": ("Powers and Indices", "Powers & Indices"),
    "B7.1.3.1": ("Fractions, Decimals and Percentages", "Fractions, Decimals & Percent"),
    "B7.1.3.2": ("Addition and Subtraction of Fractions", "Fraction Add/Subtract"),
    "B7.1.3.3": ("Multiplication and Division of Fractions", "Fraction Multiply/Divide"),
    "B7.1.4.1": ("Ratio and Proportion", "Ratio & Proportion"),
    "B7.2.1.1": ("Patterns and Relations", "Patterns & Relations"),
    "B7.2.2.1": ("Algebraic Expressions", "Algebraic Expressions"),
    "B7.2.3.1": ("Linear Equations", "Linear Equations"),
    "B7.3.1.1": ("Angle Relationships", "Angles"),
    "B7.3.1.2": ("Geometric Constructions", "Constructions"),
    "B7.3.2.1": ("Perimeter and Circumference", "Perimeter & Circumference"),
    "B7.3.2.2": ("Area of Triangles", "Triangle Area"),
    "B7.3.2.3": ("Bearings and Vectors", "Bearings & Vectors"),
    "B7.3.3.1": ("Reflection and Translation", "Reflection & Translation"),
    "B7.4.1.1": ("Data Collection and Representation", "Data Handling"),
    "B7.4.1.2": ("Mean, Median and Mode", "Central Tendency"),
    "B7.4.2.1": ("Probability of Single Events", "Single-Event Probability"),
    "B8.1.1.1": ("Standard Form and Rounding", "Standard Form & Rounding"),
    "B8.1.1.2": ("Sets, Perfect Squares and Square Roots", "Sets & Square Roots"),
    "B8.1.2.1": ("Mental Mathematics", "Mental Math"),
    "B8.1.2.2": ("Whole Number and Decimal Operations", "Whole Numbers & Decimals"),
    "B8.1.2.3": ("Laws of Indices", "Laws of Indices"),
    "B8.1.3.1": ("Fraction Operations", "Fraction Operations"),
    "B8.1.4.1": ("Ratio, Rate and Proportion", "Ratio & Proportion"),
    "B8.2.1.1": ("Linear Graphs and Gradient", "Linear Graphs"),
    "B8.2.2.1": ("Algebraic Expressions", "Algebraic Expressions"),
    "B8.2.3.1": ("Linear Inequalities", "Linear Inequalities"),
    "B8.3.1.1": ("Angles in Parallel Lines and Polygons", "Angles in Geometry"),
    "B8.3.1.2": ("Geometric Constructions and Loci", "Constructions & Loci"),
    "B8.3.2.1": ("Pythagoras, Trigonometry and Circle Area", "Pythagoras & Trigonometry"),
    "B8.3.2.2": ("Vector Addition and Subtraction", "Vector Operations"),
    "B8.3.3.1": ("Rotation", "Rotation"),
    "B8.4.1.1": ("Data Collection and Representation", "Data Representation"),
    "B8.4.1.2": ("Measures of Central Tendency and Range", "Central Tendency & Range"),
    "B8.4.2.1": ("Probability of Independent Events", "Independent Probability"),
    "B9.1.1.1": ("Place Value and Rounding", "Place Value & Rounding"),
    "B9.1.1.2": ("Rational Numbers and Sets", "Rational Numbers & Sets"),
    "B9.1.2.1": ("Mental Mathematics", "Mental Math"),
    "B9.1.2.2": ("Decimal Operations", "Decimal Operations"),
    "B9.1.2.4": ("Surds and Approximate Square Roots", "Surds"),
    "B9.1.3.1": ("Fraction Operations", "Fraction Operations"),
    "B9.1.4.1": ("Ratio, Rate and Proportion", "Ratio & Proportion"),
    "B9.2.1.1": ("Simultaneous Linear Equations", "Simultaneous Equations"),
    "B9.2.2.1": ("Formula Manipulation and Factorisation", "Formulae & Factorisation"),
    "B9.2.3.1": ("Linear Inequalities", "Linear Inequalities"),
    "B9.3.1.1": ("Angle Properties and Polygon Sums", "Angle Properties"),
    "B9.3.1.2": ("Inscribed and Circumscribed Constructions", "Geometric Constructions"),
    "B9.3.2.1": ("Surface Area of Prisms", "Surface Area of Prisms"),
    "B9.3.2.2": ("Bearings and Vectors", "Bearings & Vectors"),
    "B9.3.3.1": ("Enlargement", "Enlargement"),
    "B9.4.1.1": ("Frequency Tables and Histograms", "Tables & Histograms"),
    "B9.4.1.2": ("Descriptive Statistics", "Descriptive Statistics"),
    "B9.4.2.1": ("Probability of Dependent Events", "Dependent Probability"),
}


def update_notes_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    sheet["B3"] = (
        "The PDF defines sub-strands as the topics within each strand. "
        "Content standards state the learning aim. This revised workbook infers "
        "a concise topic title from each content-standard aim instead of repeating "
        "the objective sentence as the topic name."
    )
    sheet["B5"] = (
        "For each official content standard code, I preserved the printed code "
        "and description, then rewrote the topic title as a concise inferred topic "
        "name, with a shorter label, alternative labels, and interpretation notes."
    )
    sheet["B6"] = (
        "Where a standard covers several tightly related ideas, the topic name uses "
        "the smallest stable classroom topic title rather than the full objective wording."
    )


def update_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
    updated = 0
    for row in range(2, sheet.max_row + 1):
        code = sheet.cell(row=row, column=4).value
        if not code or code not in TOPIC_MAP:
            continue
        suggested, short = TOPIC_MAP[code]
        sheet.cell(row=row, column=7).value = suggested
        sheet.cell(row=row, column=8).value = short
        updated += 1
    return updated


def verify_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[str]:
    failures: list[str] = []
    for row in range(2, sheet.max_row + 1):
        code = sheet.cell(row=row, column=4).value
        if not code or code not in TOPIC_MAP:
            continue
        expected_suggested, expected_short = TOPIC_MAP[code]
        actual_suggested = sheet.cell(row=row, column=7).value
        actual_short = sheet.cell(row=row, column=8).value
        if actual_suggested != expected_suggested or actual_short != expected_short:
            failures.append(code)
    return failures


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.load_workbook(SOURCE_PATH)
    update_notes_sheet(workbook["Notes"])

    counts = {}
    for name in ["All_Standards", "Basic7", "Basic8", "Basic9"]:
        counts[name] = update_sheet(workbook[name])

    workbook.save(OUTPUT_PATH)

    check_wb = openpyxl.load_workbook(OUTPUT_PATH, read_only=True)
    failures = {}
    for name in ["All_Standards", "Basic7", "Basic8", "Basic9"]:
        failures[name] = verify_sheet(check_wb[name])

    print(f"Saved: {OUTPUT_PATH}")
    print("Updated counts:", counts)
    for name, codes in failures.items():
        print(f"{name} verification failures: {len(codes)}")
        if codes:
            print(", ".join(codes))


if __name__ == "__main__":
    main()
